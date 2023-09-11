# 导入所需模块.
import re
import time
import uuid
import aiohttp
import asyncio
import urllib.parse
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from selenium import webdriver
from collections import defaultdict
from selenium.webdriver.chrome.options import Options

data_dict = defaultdict(int)


# 获取视频BV号
def get_video_bv_numbers(keyword: str, bv_number_count: int):
    """
    获取与给定关键词相关的视频弹幕的 BV 号。

    参数：
    keyword (str): 搜索的关键词。
    bv_number_count (int): 需要获取的 BV 号的数量。

    返回值：
    set: 包含指定数量的与关键词相关的视频的 BV 号的集合。

    """

    # 函数代码
    print(f'开始爬取有关 {keyword} 的视频弹幕')
    options = Options()
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36 Edg/116.0.1938.69")
    # 将自动化浏览器至于后台
    options.add_argument('--headless')
    driver = webdriver.Edge(options=options)
    driver.maximize_window()
    # 添加禁用缓存的选项
    edge_options = webdriver.EdgeOptions()
    edge_options.add_argument("--disable-application-cache")
    edge_options.add_argument("--disable-gpu-shader-disk-cache")
    edge_options.add_argument("--disable-gpu-program-cache")
    edge_options.add_argument("--disable-cache")

    encoded_keyword = urllib.parse.quote(keyword)
    multiplier = 0
    res = set()
    while len(res) < bv_number_count:
        url = "https://search.bilibili.com/all?keyword=" + encoded_keyword + f'&o={24 * multiplier}'
        driver.get(url)
        driver.refresh()
        time.sleep(10)  # 保证页面加载完全
        element_content = driver.page_source
        element_content = element_content[element_content.find(r'video-list row'): element_content.find(r'to_hide_xs')]
        pattern = r'BV(.*?)/"'
        bvids = re.findall(pattern, element_content)
        res.update(bvids[: bv_number_count - len(res)])
        multiplier += 1
    driver.quit()

    return res


# 获得cid
async def get_video_cid(session, bvid: str):
    """
    异步获取指定 BV 号视频的 CID。

    参数：
    session (aiohttp.ClientSession): 异步请求会话对象。
    bvid (str): 视频的 BV 号。

    返回值：
    int: 视频的 CID，如果获取失败则返回 -1。

    """

    # 函数代码
    url = f'https://www.bilibili.com/video/{bvid}'
    async with session.get(url) as response:
        if response.status == 200:
            res = await response.text()
            res = res[res.find(r'"dynamic"'):]
            res = res[: res.find(r'"dimension"')]
            match = re.search(r'\d+', res)
            if match:  # 看是否获取到cid
                extracted_number = match.group()
                return extracted_number
        else:
            print(f"无法访问URL: {url}")
    return -1


# 获得弹幕
async def get_dm(session, cid: str):
    """
    异步获取指定 CID 的弹幕信息。

    参数：
    session (aiohttp.ClientSession): 异步请求会话对象。
    cid (str): 视频的 CID。

    返回值：
    无。

    """

    # 函数代码
    if cid == -1:
        print("cid获取失败")
        return

    url = f'https://comment.bilibili.com/{cid}.xml'
    async with session.get(url) as response:
        if response.status == 200:
            xml_data = await response.text()
            root = ET.fromstring(xml_data)

            # 统计每个弹幕出现的次数
            for element in root.findall('d'):
                data_dict[element.text] += 1
        else:
            print(f"无法访问URL: {url}")


async def process_bvid(session, bvid):
    """
    异步处理指定 BV 号的视频。

    参数：
    session (aiohttp.ClientSession): 异步请求会话对象。
    bvid (str): 视频的 BV 号。

    返回值：
    无。

    """

    # 函数代码
    cid = await get_video_cid(session, bvid)
    if cid != -1:
        await get_dm(session, cid)


# 导入EXCEL
def export_to_excel(sorted_dict: dict, excel_name: str, rows: int = -1):
    """
    将排序后的字典数据导出到 Excel 文件。

    参数：
    sorted_dict (dict): 排序后的字典，包含弹幕内容和出现次数。
    excel_name (str): 导出的 Excel 文件名。
    rows (int, 可选): 需要导出的行数，如果为 -1，则导出所有行。默认值为 -1。

    返回值：
    str: 导出的 Excel 文件名。

    """

    # 函数代码
    workbook = Workbook()
    sheet = workbook.active
    if rows == -1:
        rows = len(sorted_dict)
    rows += 1

    # 写入标题行
    sheet.cell(row=1, column=1, value='弹幕内容')
    sheet.cell(row=1, column=2, value='出现次数')

    # 写入数据
    row_num = 2
    for key, value in sorted_dict.items():
        if row_num > rows:
            break
        sheet.cell(row=row_num, column=1, value=key)
        sheet.cell(row=row_num, column=2, value=value)
        row_num += 1

    # 通过uuid防止文件名重复
    filename = f'{excel_name}_{uuid.uuid4()}.xlsx'
    workbook.save(filename)

    return filename


async def generate_excel():
    """
    异步生成包含指定关键词搜索结果的 Excel 文件。

    参数：
    无。

    返回值：
    str: 生成的 Excel 文件名。

    """

    # 函数代码
    print("输入搜索关键词:")
    keyword = input()
    print("输入要搜索的视频数量:")
    video_num = input()
    print("输入要导出的弹幕行数:")
    rows = input()
    print(f"等待时间约为: {((int(video_num) // 24) + 1) * 12 + 58} s")
    bvids = ['BV' + bvid for bvid in get_video_bv_numbers(str(keyword), int(video_num))]
    print(f'本次爬取视频的bvid为: {bvids}')
    print('开始获取弹幕')
    async with aiohttp.ClientSession() as session:
        tasks = [asyncio.create_task(process_bvid(session, bvid)) for bvid in bvids]
        await asyncio.gather(*tasks)
    sorted_dict = dict(sorted(data_dict.items(), key=lambda x: x[1], reverse=True))
    res = export_to_excel(sorted_dict, str(keyword), int(rows))
    return res


async def main():
    await generate_excel()


if __name__ == '__main__':
    asyncio.get_event_loop().run_until_complete(main())
