# 导入所需模块.
import pkuseg
import numpy as np
import matplotlib.font_manager as fm
from PIL import Image
from wordcloud import WordCloud
from openpyxl import load_workbook
from collections import defaultdict


def generate_wordcloud(excel_name: str, keywords: dict, stopwords: list, img_path: str) -> None:
    """
    生成词云图。

    参数：
    excel_name (str): Excel 文件名，包含文本数据。
    keywords (dict): 自定义词典，用于分词时的关键词匹配。
    stopwords (list): 停用词列表，用于过滤分词结果。

    返回值：
    无。

    """

    # 函数代码
    # 从 Excel 文件读取文本数据
    text_data = {}
    try:
        workbook = load_workbook(excel_name)
    except FileNotFoundError:
        print(f"文件 '{excel_name}' 不存在")
        return
    except Exception as e:
        print(f"读取文件 '{excel_name}' 时发生错误: {str(e)}")
        return
    sheet = workbook.active

    keyword_freq = {}
    header_skipped = False  # 标记是否已跳过表头行

    for row in sheet.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue  # 跳过表头行

        sentence = str(row[0])
        frequency = int(row[1])

        text_data[sentence] = frequency

    seg = pkuseg.pkuseg(user_dict=keywords)
    seg_text_data = defaultdict(int)

    for sentence, frequency in text_data.items():
        seg_sentence = seg.cut(sentence)
        filtered_sentence = [word for word in seg_sentence if word not in stopwords]
        seg_sentence_str = ' '.join(filtered_sentence)
        seg_text_data[seg_sentence_str] += frequency

    mask_image = np.array(Image.open(img_path))
    font_path = fm.findfont(fm.FontProperties(family="SimSun"))
    # 生成词云图
    wc = WordCloud(font_path=font_path, width=800, height=400, mask=mask_image,
                   background_color='white').generate_from_frequencies(seg_text_data)

    # 保存词云图
    image = wc.to_image()
    image.save(f"{excel_name[:excel_name.rfind('.xlsx')]}.png")

    # 显示词云图
    image.show()


def main():
    print("输入EXCEL文件名:")
    excel_name = input()
    generate_wordcloud(str(excel_name), {}, [], r"images/dove-35401_1920.png")


if __name__ == '__main__':
    main()
